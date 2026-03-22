import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 새 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "전기요금청구서"

# 칼럼 타이틀
headers = ["년월", "사용기간", "사용량(kWh)", "청구금액"]

# 헤더 스타일 설정
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

# 헤더 작성
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 칼럼 너비 조정
column_widths = {
    'A': 12,  # 년월
    'B': 25,  # 사용기간
    'C': 15,  # 사용량(kWh)
    'D': 15,  # 청구금액
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 행 높이 조정
ws.row_dimensions[1].height = 25

# 파일 저장
filename = "전기요금청구서.xlsx"
wb.save(filename)
print(f"✅ {filename} 파일이 생성되었습니다.")
